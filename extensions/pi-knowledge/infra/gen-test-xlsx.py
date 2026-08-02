"""生成测试用噪声 Excel：企业系统导出的销售报表（中文）"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = r"D:\knowledge-vault\test-dirty"
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb = Workbook()

# ============================================================
# Sheet 1: 销售数据（主数据 sheet，但前面有大量导出元信息）
# ============================================================
ws = wb.active
ws.title = "销售数据"

# --- 导出元信息噪声（前 18 行） ---
meta_font = Font(size=9, color="666666")
title_font = Font(size=14, bold=True)
header_font = Font(size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

ws.merge_cells("A1:H1")
ws["A1"] = "星辰科技有限公司 - ERP数据导出报表"
ws["A1"].font = title_font

meta_rows = [
    ("报表编号：", "RPT-2024-06-0892"),
    ("报表名称：", "2024年Q2区域销售明细表"),
    ("导出系统：", "星辰ERP V8.2 (SAP兼容模式)"),
    ("导出时间：", "2024-07-01 08:15:32 CST"),
    ("导出人：", "系统管理员 (admin)"),
    ("数据范围：", "2024-04-01 至 2024-06-30"),
    ("筛选条件：", "区域=全部 | 产品线=全部 | 客户等级>=B | 含退货=否"),
    ("排序方式：", "区域 ASC → 销售额 DESC"),
    ("币种：", "人民币(CNY) | 汇率基准: 2024-06-30 央行中间价"),
    ("数据条数：", "共 156 条记录（本页显示前 30 条）"),
    ("分页：", "第 1 页 / 共 6 页"),
    ("", ""),
    ("【注意】本报表由系统自动生成，请勿手工修改。如有疑问请联系数据分析组。", ""),
    ("【密级】内部机密 - 仅限部门经理及以上级别查阅", ""),
    ("", ""),
]
for i, (label, value) in enumerate(meta_rows, start=2):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = meta_font
    ws[f"B{i}"] = value
    ws[f"B{i}"].font = meta_font

# --- 空行分隔 ---
data_start = 18

# --- 表头（第 18 行）---
headers = [
    "序号", "区域", "省份", "客户名称", "客户等级", "产品线",
    "产品名称", "规格型号", "销售数量", "单价(元)", "金额(元)",
    "折扣率", "实收金额(元)", "销售日期", "销售员", "备注"
]
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=data_start, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# --- 数据行（30 条模拟数据）---
import random
random.seed(42)

regions = ["华东", "华南", "华北", "西南", "华中"]
provinces = {
    "华东": ["江苏", "浙江", "上海", "安徽"],
    "华南": ["广东", "福建", "广西"],
    "华北": ["北京", "天津", "河北", "山东"],
    "西南": ["四川", "重庆", "云南"],
    "华中": ["湖北", "湖南", "河南"],
}
customers = [
    ("华为技术有限公司", "A"), ("阿里巴巴集团", "A"), ("腾讯科技", "A"),
    ("字节跳动", "B"), ("京东集团", "B"), ("美团点评", "B"),
    ("小米科技", "B"), ("网易公司", "C"), ("百度在线", "B"),
    ("中兴通讯", "C"), ("联想集团", "C"), ("大疆创新", "B"),
    ("比亚迪股份", "A"), ("宁德时代", "A"), ("格力电器", "C"),
]
products = [
    ("智能网关", "XG-3000", "企业版"),
    ("边缘计算节点", "EC-500", "标准版"),
    ("数据传感器", "DS-100T", "工业级"),
    ("云平台许可证", "CP-ENT", "年度订阅"),
    ("AI推理卡", "AI-R200", "旗舰版"),
]
salespersons = ["张伟", "李娜", "王强", "刘芳", "陈明", "赵丽", "孙磊"]

for i in range(30):
    row = data_start + 1 + i
    region = random.choice(regions)
    province = random.choice(provinces[region])
    cust_name, cust_level = random.choice(customers)
    prod_name, prod_model, prod_ver = random.choice(products)
    qty = random.randint(5, 500)
    price = random.choice([2800, 5600, 1200, 45000, 8900])
    amount = qty * price
    discount = random.choice([1.0, 0.95, 0.9, 0.85, 0.8])
    actual = int(amount * discount)
    month = random.choice([4, 5, 6])
    day = random.randint(1, 28)
    date_str = f"2024-{month:02d}-{day:02d}"
    salesperson = random.choice(salespersons)
    note = random.choice(["", "", "", "大客户专项折扣", "季度冲量", "新客首单", ""])

    values = [
        i + 1, region, province, cust_name, cust_level, prod_name,
        f"{prod_name} {prod_ver}", prod_model, qty, price, amount,
        f"{int(discount*100)}%", actual, date_str, salesperson, note
    ]
    for col, val in enumerate(values, start=1):
        ws.cell(row=row, column=col, value=val)

# --- 汇总行噪声 ---
summary_row = data_start + 31
ws.cell(row=summary_row, column=1, value="").font = meta_font
ws.cell(row=summary_row + 1, column=1, value="本页小计：").font = Font(bold=True)
ws.cell(row=summary_row + 1, column=11, value="¥ 28,456,800")
ws.cell(row=summary_row + 1, column=13, value="¥ 25,611,120")
ws.cell(row=summary_row + 2, column=1, value="全部合计：").font = Font(bold=True)
ws.cell(row=summary_row + 2, column=11, value="¥ 156,782,400")
ws.cell(row=summary_row + 2, column=13, value="¥ 141,104,160")
ws.cell(row=summary_row + 3, column=1, value="").font = meta_font
ws.cell(row=summary_row + 4, column=1, value="制表人：数据分析组 周敏 | 审核人：财务总监 马建国 | 日期：2024-07-01").font = meta_font
ws.cell(row=summary_row + 5, column=1, value="本报表数据来源于ERP系统，最终解释权归数据分析组所有。").font = meta_font
ws.cell(row=summary_row + 6, column=1, value="如需完整数据请联系 data-analysis@xingchen-tech.com 或内线 8023。").font = meta_font

# 设置列宽
for col in range(1, 17):
    ws.column_dimensions[get_column_letter(col)].width = 14

# ============================================================
# Sheet 2: 使用说明（纯噪声 sheet）
# ============================================================
ws2 = wb.create_sheet("使用说明")
instructions = [
    "【报表使用说明】",
    "",
    "1. 本报表由星辰ERP系统自动导出，数据截止时间为导出时刻。",
    "2. 报表中的数据已经过以下处理：",
    "   - 退货订单已剔除（如需含退货版本，请在导出时勾选）",
    "   - 金额已按折扣率计算实收金额",
    "   - 跨币种订单已按导出日汇率折算为人民币",
    "3. 本报表每页显示30条记录，完整数据共156条（6页）。",
    "   如需完整数据，请在ERP中重新导出或联系数据分析组。",
    "4. 客户等级说明：",
    "   A级：年采购额>500万 | B级：100-500万 | C级：<100万",
    "5. 产品线编码规则：",
    "   XG=智能网关 | EC=边缘计算 | DS=数据传感器 | CP=云平台 | AI=AI推理",
    "",
    "【版本历史】",
    "V1.0 (2024-04) - 初版，仅含华东区域",
    "V1.1 (2024-05) - 扩展至全部区域",
    "V2.0 (2024-06) - 增加折扣率和实收金额列",
    "V2.1 (2024-07) - 增加客户等级和产品规格列",
    "",
    "【联系方式】",
    "数据分析组：data-analysis@xingchen-tech.com | 内线 8023",
    "ERP运维组：erp-support@xingchen-tech.com | 内线 8001",
    "IT服务台：it-helpdesk@xingchen-tech.com | 内线 9000",
    "",
    "【免责声明】",
    "本报表仅供内部管理决策参考，不构成任何财务审计依据。",
    "数据以ERP系统实时查询结果为准，本报表可能存在导出时点差异。",
    "未经授权不得将本报表发送给外部人员或用于非公司目的。",
    "",
    "星辰科技有限公司 数据分析组",
    "文档编号：RPT-2024-06-0892 | 密级：内部机密",
]
for i, line in enumerate(instructions, start=1):
    ws2[f"A{i}"] = line
    ws2[f"A{i}"].font = Font(size=9)
ws2.column_dimensions["A"].width = 80

# ============================================================
# Sheet 3: 模板(勿动)（占位模板 sheet）
# ============================================================
ws3 = wb.create_sheet("模板(勿动)")
ws3["A1"] = "⚠️ 本Sheet为系统模板，请勿修改或删除！修改将导致报表导出异常。"
ws3["A1"].font = Font(size=11, bold=True, color="FF0000")

template_headers = ["${SEQ}", "${REGION}", "${PROVINCE}", "${CUSTOMER}", "${LEVEL}",
                    "${PRODUCT}", "${FULL_NAME}", "${MODEL}", "${QTY}", "${PRICE}",
                    "${AMOUNT}", "${DISCOUNT}", "${ACTUAL}", "${DATE}", "${SALES}", "${NOTE}"]
for col, h in enumerate(template_headers, start=1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = Font(size=9, color="999999")
    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# 占位数据
for row in range(4, 8):
    for col in range(1, 17):
        ws3.cell(row=row, column=col, value=f"{{{{FIELD_{col}}}}}")
        ws3.cell(row=row, column=col).font = Font(size=9, color="CCCCCC")

ws3["A10"] = "模板版本: 2.1 | 最后修改: 2024-06-15 by admin | 关联报表: RPT-2024-*"
ws3["A10"].font = Font(size=8, color="999999")
ws3["A11"] = "如需修改模板请联系ERP运维组(erp-support@xingchen-tech.com)，自行修改后果自负。"
ws3["A11"].font = Font(size=8, color="999999")

# ============================================================
# Sheet 4: 数据透视表（另一个有数据的 sheet）
# ============================================================
ws4 = wb.create_sheet("区域汇总")
ws4["A1"] = "2024年Q2区域销售汇总（自动生成，请勿修改）"
ws4["A1"].font = Font(size=11, bold=True)
ws4["A2"] = "导出时间: 2024-07-01 08:15:32 | 数据来源: 销售数据Sheet | 汇率: 1.0"
ws4["A2"].font = Font(size=8, color="666666")

pivot_headers = ["区域", "订单数", "客户数", "销售总额(元)", "实收总额(元)", "平均折扣", "同比增长"]
for col, h in enumerate(pivot_headers, start=1):
    cell = ws4.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill

pivot_data = [
    ("华东", 52, 18, 52_340_000, 47_106_000, "90%", "+12.3%"),
    ("华南", 38, 14, 38_120_000, 34_308_000, "90%", "+8.7%"),
    ("华北", 31, 12, 31_560_000, 28_404_000, "90%", "+15.1%"),
    ("西南", 20, 8, 20_450_000, 18_405_000, "90%", "+5.2%"),
    ("华中", 15, 7, 14_312_400, 12_881_160, "90%", "+3.8%"),
    ("合计", 156, 45, 156_782_400, 141_104_160, "90%", "+10.2%"),
]
for i, row_data in enumerate(pivot_data, start=5):
    for col, val in enumerate(row_data, start=1):
        cell = ws4.cell(row=i, column=col, value=val)
        if i == 10:  # 合计行加粗
            cell.font = Font(bold=True)

ws4["A12"] = "注：同比增长基于2023年Q2同口径数据计算。"
ws4["A12"].font = Font(size=8, color="666666")
ws4["A13"] = "制表: 周敏 | 审核: 马建国 | 批准: 赵鹏飞(CTO)"
ws4["A13"].font = Font(size=8, color="666666")

for col in range(1, 8):
    ws4.column_dimensions[get_column_letter(col)].width = 16

# ============================================================
# 保存
# ============================================================
filepath = os.path.join(OUTPUT_DIR, "2024Q2区域销售明细表.xlsx")
wb.save(filepath)
print(f"已生成: {filepath}")
print(f"文件大小: {os.path.getsize(filepath) / 1024:.1f} KB")
print(f"Sheets: {wb.sheetnames}")
