#!/usr/bin/env python3
"""Scaffold the raw/ corpus for multi-modal RAG benchmarking.

Generates the English binaries (1 PDF, 2 Excels, 5 CSVs, 2 generated PNGs, 5
stock JPGs) plus the multi-language (Japanese/Chinese) binaries (2 CJK PDFs,
1 Excel, 5 CSVs, 1 PNG, 2 JPGs). Markdown files are written separately via the
Write tool so cross-links stay reviewable in the diff.

Run (from repo root):
    uv run --with reportlab,openpyxl,Pillow,matplotlib,PyMuPDF \\
        scripts/generation/scaffold-personal.py            # all binaries
    uv run --with reportlab,openpyxl,Pillow,PyMuPDF \\
        scripts/generation/scaffold-personal.py --only cjk # CJK binaries only
"""

from __future__ import annotations

import argparse
import csv
import math
import random
import sys
import urllib.error
import urllib.request
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]  # scripts/generation/ -> repo root
P = ROOT / "raw"  # corpus lives directly under raw/ (the historical raw/personal/ prefix was removed)

RNG = random.Random(42)


# ---------------------------------------------------------------------------
# Paper content (Dr. Wren Adachi, coral resilience)
# ---------------------------------------------------------------------------

PAPER = {
    "title": ("Thermal Forcing and Coral Bleaching Resilience in the "
              "Ryukyu Archipelago: A Multi-Site Survey, 2020–2024"),
    "authors": ["Wren Adachi", "Hajime Tanaka", "Maria Santos", "Kenta Higa"],
    "affils": [
        "Institute of Marine Sciences, UC Santa Cruz",
        "Sesoko Station, Tropical Biosphere Research Center, University of the Ryukyus",
        "Okinawa Churaumi Aquarium Research Division",
    ],
    "abstract": (
        "Coral bleaching events have increased in frequency and severity across the "
        "Indo-Pacific, yet resilience patterns remain undersampled at mid-latitude reef "
        "systems. Between 2020 and 2024 we conducted quarterly surveys at 12 reef sites "
        "spanning 24°N–27°N in the Ryukyu Archipelago, combining in-situ sea-surface "
        "temperature (SST) loggers, quadrat-based bleaching indices, and symbiont "
        "community profiling. Bleaching index scales nonlinearly with cumulative thermal "
        "anomaly (degree heating weeks, DHW), with a critical threshold near DHW = 8.4 "
        "above which mean bleaching exceeds 60%. Two northern sites (Cape Hedo, Sesoko "
        "West) maintained <20% bleaching during the 2023 Pacific marine heatwave despite "
        "DHW > 10, suggesting local thermal adaptation or symbiont community buffering. "
        "These refugia are priority candidates for Pacific-wide conservation corridor "
        "planning under RCP 4.5 and RCP 8.5 trajectories. We discuss implications for "
        "managed relocation of resilient genotypes and identify three unresolved questions "
        "concerning the role of background microbial community structure in buffering acute "
        "thermal stress."
    ),
    "keywords": ("coral bleaching, Ryukyu Archipelago, thermal resilience, "
                 "degree heating weeks, symbiont shuffling, climate refugia"),
    "sections": [
        ("1. Introduction", [
            "The global coral crisis has accelerated through the past three decades, with "
            "three mass bleaching events on the Great Barrier Reef alone since 2016 (Hughes "
            "et al. 2017; Dietzel et al. 2020). While tropical reef systems have received "
            "the bulk of monitoring attention, mid-latitude reefs in the West Pacific—long "
            "hypothesised to serve as climate refugia—remain sparsely sampled at multi-year "
            "resolution.",

            "The Ryukyu Archipelago extends roughly 1,000 km from Kyushu to Taiwan, "
            "straddling the Kuroshio Current. Its reefs experience strong seasonal SST "
            "amplitude (14–30 °C) and episodic heat anomalies driven by the Western Pacific "
            "Warm Pool teleconnection. Despite their biogeographic importance, no "
            "standardised multi-site bleaching time series at quarterly cadence has been "
            "published for the archipelago as a whole.",

            "Here we report on four years (2020 Q1 – 2024 Q4, 16 quarters) of paired "
            "physical and biological observations across 12 sites, with three questions: "
            "(i) what is the shape of the bleaching–thermal forcing relationship; (ii) are "
            "there sites that consistently depart from the regional norm; and (iii) do such "
            "resilience outliers share measurable predictors (depth, exposure, symbiont "
            "community)?",
        ]),
        ("2. Methods", [
            "2.1 Study sites. Twelve sites were selected across four island groups "
            "(Yonaguni, Ishigaki, Okinawa-honto, Amami-oshima) spanning 24.3°N–27.0°N "
            "(Table S1). Site selection balanced forereef vs. backreef, exposure (NE, SW, "
            "leeward), and depth (5–18 m). Full coordinates and survey design are listed in "
            "the supplementary dive-sites.csv.",

            "2.2 Temperature logging. Onset HOBO U22-001 loggers were deployed at each "
            "site at 8 m depth, recording at 30-minute intervals. Loggers were swapped and "
            "calibrated annually against an NIST-traceable reference. Degree heating weeks "
            "(DHW) were computed following NOAA Coral Reef Watch v3.1, using the site-specific "
            "climatology from 2010–2019.",

            "2.3 Bleaching index. At each site, five permanent 25 m transects were "
            "photographed quarterly using a Canon R5 in a Nauticam housing with dual "
            "Inon Z-330 strobes. Quadrats (1 m², n = 50 per transect) were scored in Adobe "
            "Lightroom against the standardised Chasing Coral colour reference card. "
            "Bleaching index was defined as the percentage of live coral colonies "
            "exhibiting ≥2-grade colour loss (1–6 scale).",

            "2.4 Symbiont community. At a subset of six sites, tissue biopsies (n = 12 "
            "per site per quarter) were preserved in 95% EtOH and shipped to UC Santa Cruz "
            "for ITS2 metabarcoding (primers sym_VAR_5.8S2/sym_VAR_REV). Operational "
            "taxonomic units were clustered at 97% identity against the GeoSymbio database.",

            "2.5 Statistical analysis. We fit generalised additive mixed models (GAMMs) "
            "with site as a random effect and DHW, exposure, and depth as fixed effects, "
            "using a beta regression link appropriate for proportional data. All analyses "
            "were run in R 4.3 with the mgcv and glmmTMB packages; code is archived at "
            "github.com/wadachi/ryukyu-resilience.",
        ]),
        ("3. Results", [
            "3.1 Thermal forcing. SST across the 12 sites ranged from 14.2 °C (Cape "
            "Hedo, Feb 2021) to 31.8 °C (Yonaguni, Aug 2023). Cumulative DHW exceeded the "
            "NOAA Alert Level 2 threshold (DHW ≥ 8) at 9 of 12 sites during the 2023 "
            "heatwave, peaking at DHW = 12.4 at Ishigaki South.",

            "3.2 Bleaching–DHW relationship. Mean bleaching index scaled nonlinearly with "
            "DHW (Fig. 2). A piecewise regression identified a breakpoint at DHW = 8.36 "
            "(95% CI 7.91–8.82), above which bleaching increased at 4.1× the baseline rate. "
            "Above DHW = 10, mean bleaching exceeded 60% across all sites pooled.",

            "3.3 Resilience outliers. Two northern forereef sites—Cape Hedo and Sesoko "
            "West—maintained mean bleaching < 20% throughout 2023 despite DHW > 10 at both. "
            "Mixed-effects models attribute 38% of this residual variance to symbiont "
            "community evenness (Shannon H′, p < 0.001) and a further 17% to mean colony "
            "depth (p = 0.004).",

            "3.4 Symbiont turnover. Six sites showed the expected Acropora–Durusdinium "
            "shuffle during peak DHW. The two resilience outliers instead hosted "
            "Cladocopium C3k dominance persistently, with no detectable shuffle. This "
            "divergence is the strongest candidate mechanism for the observed resilience.",
        ]),
        ("4. Discussion", [
            "Our four-year time series confirms that bleaching in the Ryukyu Archipelago "
            "scales nonlinearly with cumulative thermal anomaly, consistent with regional "
            "patterns from the Great Barrier Reef and the Maldives. The breakpoint at "
            "DHW ≈ 8.4 is statistically indistinguishable from the NOAA Alert Level 2 "
            "threshold, lending operational support to existing watch products.",

            "The two resilience outliers are, to our knowledge, the first documented cases "
            "of sustained coral resistance to DHW > 10 in the Ryukyu Archipelago. Their "
            "shared signature—persistent Cladocopium C3k dominance, deeper mean colony "
            "depth, NE exposure—is consistent with the emerging 'background symbiont "
            "community' hypothesis (Boulotte et al. 2016; Cunning et al. 2018), which "
            "posits that thermal tolerance is set by standing symbiont diversity rather "
            "than post-stress shuffling.",

            "Three caveats apply. First, our survey does not resolve post-bleaching "
            "mortality; resilient sites may still suffer sublethal fitness costs. Second, "
            "symbiont community is itself a moving target under RCP 8.5 warming trajectories. "
            "Third, the geographic breadth of resilience refugia may shrink faster than "
            "models predict if heatwaves outpace the rate of symbiont community turnover.",

            "We close with a call for coordinated trans-pacific monitoring of symbiont "
            "community structure, alongside the existing temperature watch network. Without "
            "it, the conservation corridors we propose in Section 5 rest on inference from "
            "twelve sites—an unstable foundation for Pacific-wide planning.",
        ]),
        ("5. Conclusion", [
            "Coral reefs of the Ryukyu Archipelago display a robust nonlinear response to "
            "thermal forcing, with two sites exhibiting resilient outliers consistent with "
            "the background-symbiont-community hypothesis. These refugia merit immediate "
            "conservation attention and long-term monitoring under projected climate "
            "trajectories.",
        ]),
    ],
    "table_rows": [
        # site, mean_SST_C, mean_DHW, mean_bleaching_pct, dominant_symiont, refugia_flag
        ["Yonaguni West",     27.4, 9.1,  62.3, "Durusdinium",   "no"],
        ["Ishigaki South",    28.1, 11.4, 78.1, "Durusdinium",   "no"],
        ["Kabira North",      27.9, 9.8,  55.6, "Cladocopium C3",  "no"],
        ["Sekisei Lagoon",    28.4, 10.9, 71.4, "Durusdinium",   "no"],
        ["Cape Hedo",         26.3, 10.2, 18.1, "Cladocopium C3k", "yes"],
        ["Sesoko West",       26.5, 10.5, 17.4, "Cladocopium C3k", "yes"],
        ["Maeda Point",       26.8, 9.0,  44.7, "Cladocopium C3",  "no"],
        ["Onna Village East", 26.9, 8.7,  39.2, "Cladocopium C3",  "no"],
        ["Kerama Inner",      27.1, 8.4,  36.8, "Cladocopium C3",  "no"],
        ["Kerama Outer",      27.0, 9.2,  48.5, "Cladocopium C3",  "no"],
        ["Amami North",       26.6, 8.1,  31.9, "Cladocopium C3",  "no"],
        ["Amami South",       26.8, 8.6,  38.5, "Cladocopium C3",  "no"],
    ],
    "references": [
        "Boulotte, N.M., Dalton, S.J., Carroll, A.G., Harrison, P.L., Putnam, H.M., "
        "Fabricius, K.E., Bollati, E., Madin, J.S., & Buerger, P.T. (2016). Exploring the "
        "Symbiodinium rare biosphere provides evidence for symbiont diversity in reef-building "
        "corals. Molecular Ecology, 25(17), 4501–4517.",

        "Cunning, R., Silverstein, R.N., & Baker, A.C. (2018). Warming patterns associated "
        "with symbiont shuffling in a model cnidarian. ISME Journal, 12(2), 422–434.",

        "Dietzel, A., Bode, M., Connolly, S.R., & Hughes, T.P. (2020). Long-term shifts in "
        "mass coral bleaching mortality. Proceedings of the Royal Society B, 287, 2020–2087.",

        "Hughes, T.P., Kerry, J.T., Álvarez-Noriega, M., Álvarez-Romero, J.G., Anderson, "
        "K.D., Baird, A.H., Babcock, R.C., Beger, M., Bellwood, D.R., Berkelmans, R., et al. "
        "(2017). Global warming and recurrent mass bleaching of corals. Nature, 543, 373–377.",

        "NOAA Coral Reef Watch (2023). Daily global 5-km satellite coral bleaching heat "
        "product suite, version 3.1. NOAA National Environmental Satellite, Data, and "
        "Information Service.",

        "Putnam, H.M., Davidson, J.M., & Gates, R.D. (2016). Ocean acidification influences "
        "host microbiome: a perspective on the role of symbiont community in coral health. "
        "Frontiers in Microbiology, 7, 401.",

        "Van Oppen, M.J.H., Oliver, J.K., Putnam, H.M., & Gates, R.D. (2015). Building coral "
        "reef resilience through assisted evolution. Proceedings of the National Academy of "
        "Sciences, 112(8), 2307–2313.",
    ],
}


# ---------------------------------------------------------------------------
# Study site data
# ---------------------------------------------------------------------------

SITES = [
    # site_id, name, island_group, lat, lon, depth_m, exposure, reef_type
    ("RYK-01", "Yonaguni West",     "Yonaguni",       24.452, 122.746, 12, "SW", "forereef"),
    ("RYK-02", "Ishigaki South",    "Ishigaki",       24.331, 124.156, 10, "S",  "forereef"),
    ("RYK-03", "Kabira North",      "Ishigaki",       24.498, 124.064, 8,  "N",  "backreef"),
    ("RYK-04", "Sekisei Lagoon",    "Ishigaki",       24.220, 124.090, 6,  "S",  "lagoon"),
    ("RYK-05", "Cape Hedo",         "Okinawa-honto",  26.864, 128.146, 14, "NE", "forereef"),
    ("RYK-06", "Sesoko West",       "Okinawa-honto",  26.632, 127.864, 15, "W",  "forereef"),
    ("RYK-07", "Maeda Point",       "Okinawa-honto",  26.444, 127.742, 8,  "W",  "forereef"),
    ("RYK-08", "Onna Village East", "Okinawa-honto",  26.510, 127.912, 10, "E",  "forereef"),
    ("RYK-09", "Kerama Inner",      "Kerama",         26.198, 127.310, 7,  "W",  "lagoon"),
    ("RYK-10", "Kerama Outer",      "Kerama",         26.124, 127.271, 18, "SW", "forereef"),
    ("RYK-11", "Amami North",       "Amami-oshima",   28.432, 129.624, 11, "N",  "forereef"),
    ("RYK-12", "Amami South",       "Amami-oshima",   28.166, 129.438, 9,  "S",  "backreef"),
]

SPECIES = [
    "Acropora digitifera",
    "Acropora muricata",
    "Porites lutea",
    "Pocillopora verrucosa",
    "Montipora aequituberculata",
    "Favites chinensis",
    "Platygyra daedalea",
    "Goniastrea retiformis",
]

# ---------------------------------------------------------------------------
# PDF writer (reportlab)
# ---------------------------------------------------------------------------

def write_pdf(target: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle, KeepTogether
    )

    target.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(target),
        pagesize=LETTER,
        leftMargin=1.0 * inch, rightMargin=1.0 * inch,
        topMargin=1.0 * inch, bottomMargin=1.0 * inch,
        title=PAPER["title"],
        author=", ".join(PAPER["authors"]),
    )
    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title2", parent=base["Title"], fontSize=18, leading=22, spaceAfter=14)
    h1 = ParagraphStyle("H1", parent=base["Heading1"], fontSize=13, leading=17,
                        spaceBefore=14, spaceAfter=8)
    body = ParagraphStyle("Body2", parent=base["BodyText"], fontSize=10.5,
                          leading=14, spaceAfter=8, alignment=4)  # justify
    small = ParagraphStyle("Small", parent=base["BodyText"], fontSize=9, leading=12)
    abstr = ParagraphStyle("Abstract", parent=body, leftIndent=18, rightIndent=18,
                           fontSize=9.5, leading=13)

    flow = []
    flow.append(Paragraph(PAPER["title"], title_style))
    flow.append(Paragraph(", ".join(PAPER["authors"]), body))
    flow.append(Paragraph("<br/>".join(PAPER["affils"]), small))
    flow.append(Spacer(1, 10))
    flow.append(Paragraph("<b>Abstract.</b> " + PAPER["abstract"], abstr))
    flow.append(Spacer(1, 6))
    flow.append(Paragraph("<b>Keywords:</b> " + PAPER["keywords"], small))
    flow.append(PageBreak())

    for title, paragraphs in PAPER["sections"]:
        flow.append(Paragraph(title, h1))
        for p in paragraphs:
            flow.append(Paragraph(p, body))
        flow.append(Spacer(1, 6))

    # Table 1
    flow.append(Paragraph("Table 1. Site-level summary (2020–2024 mean).", h1))
    tbl_data = [["Site", "Mean SST (°C)", "Mean DHW", "Bleaching (%)",
                 "Dominant symbiont", "Refugia?"]]
    tbl_data.extend(PAPER["table_rows"])
    tbl = Table(tbl_data, repeatRows=1, colWidths=[1.3*inch, 0.95*inch, 0.7*inch,
                                                   0.95*inch, 1.4*inch, 0.6*inch])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a3b5c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.white, colors.HexColor("#f0f4f8")]),
        ("BACKGROUND", (0, 5), (-1, 6), colors.HexColor("#fff4e0")),  # refugia highlight
    ]))
    flow.append(tbl)
    flow.append(Spacer(1, 8))
    flow.append(Paragraph(
        "Sites highlighted in amber are the two resilience refugia discussed in §3.3.", small))

    # References
    flow.append(PageBreak())
    flow.append(Paragraph("References", h1))
    for ref in PAPER["references"]:
        flow.append(Paragraph(ref, small))
        flow.append(Spacer(1, 4))

    doc.build(flow)


# ---------------------------------------------------------------------------
# Excel writers (openpyxl)
# ---------------------------------------------------------------------------

def write_excels(base: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    (base / "research").mkdir(parents=True, exist_ok=True)
    (base / "fermentation").mkdir(parents=True, exist_ok=True)

    header_fill = PatternFill("solid", fgColor="1A3B5C")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    refugia_fill = PatternFill("solid", fgColor="FFE8B0")
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    def style_sheet(ws, header_row=1):
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 32)

    # ---- bleaching-survey-2024.xlsx ----
    wb = Workbook()

    # Sheet 1: Sites
    ws = wb.active
    ws.title = "Sites"
    headers = ["site_id", "name", "island_group", "latitude", "longitude",
               "depth_m", "exposure", "reef_type"]
    ws.append(headers)
    for row in SITES:
        ws.append(row)
    style_sheet(ws)

    # Sheet 2: Species cover (% per site)
    ws = wb.create_sheet("SpeciesCover_pct")
    ws.append(["species"] + [s[0] for s in SITES])
    for sp in SPECIES:
        ws.append([sp] + [round(RNG.uniform(0.5, 32.5), 1) for _ in SITES])
    style_sheet(ws)

    # Sheet 3: Monthly SST (°C) — Jan 2024 across sites
    ws = wb.create_sheet("SST_2024_monthly")
    ws.append(["month"] + [s[0] for s in SITES])
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_base = [21.0, 21.3, 22.4, 24.1, 26.0, 28.0, 29.4, 29.7, 28.6, 26.8, 24.2, 21.8]
    # Northern sites are ~1-2°C cooler, southern warmer
    site_offset = {s[0]: (26.5 - s[3]) * -0.8 for s in SITES}  # lat-based offset
    for i, m in enumerate(months):
        row = [m]
        for s in SITES:
            base_t = monthly_base[i] + site_offset[s[0]] + RNG.uniform(-0.3, 0.3)
            row.append(round(base_t, 2))
        ws.append(row)
    style_sheet(ws)

    # Sheet 4: Bleaching (%) per quarter, 2020–2024 (20 quarters)
    ws = wb.create_sheet("Bleaching_quarterly_pct")
    ws.append(["quarter"] + [s[0] for s in SITES])
    quarters = [f"{y}Q{q}" for y in range(2020, 2025) for q in range(1, 5)]
    refugia_ids = {"RYK-05", "RYK-06"}
    for q in quarters:
        year = int(q[:4])
        # 2023 was the big heatwave
        heatwave_factor = {2020: 0.6, 2021: 0.7, 2022: 0.8, 2023: 1.6, 2024: 0.9}[year]
        row = [q]
        for s in SITES:
            site_id = s[0]
            # Latitude effect: southern sites warmer so more bleaching
            lat_factor = (27.0 - s[3]) * 4
            if site_id in refugia_ids:
                bleached = RNG.uniform(2, 18)
            else:
                bleached = RNG.uniform(15, 55) + lat_factor
            row.append(round(min(95, max(0, bleached * heatwave_factor)), 1))
        ws.append(row)
    style_sheet(ws)
    # Highlight refugia columns
    for col_idx, s in enumerate(SITES, start=2):
        if s[0] in refugia_ids:
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).fill = refugia_fill

    wb.save(base / "research" / "bleaching-survey-2024.xlsx")

    # ---- fermentation/batch-log.xlsx ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"
    ws.append(["batch_id", "type", "start_date", "duration_days", "vessel",
               "temp_c", "start_ph", "end_ph", "gravity_start", "gravity_end",
               "rating_1to5", "notes"])
    types = [
        ("kombucha",   (22, 26), 7, 14, "1-gal jar"),
        ("koji",       (28, 32), 10, 18, "cedar tray"),
        ("hot_sauce",  (22, 26), 7, 21, "swing-top bottle"),
        ("kimchi",     (18, 22), 5, 14, "clamp jar"),
        ("miso",       (28, 32), 60, 365, "ceramic crock"),
    ]
    start = date(2023, 6, 1)
    for i in range(1, 21):
        t = RNG.choice(types)
        name, (tlo, thi), dur_min, dur_max, vessel = t
        d = start + timedelta(days=RNG.randint(0, 540))
        duration = RNG.randint(dur_min, dur_max)
        temp_c = round(RNG.uniform(tlo, thi), 1)
        start_ph = round(RNG.uniform(3.0 if name == "kombucha" else 5.5,
                                      4.5 if name == "kombucha" else 6.5), 2)
        end_ph = round(max(2.8, start_ph - RNG.uniform(0.3, 1.8)), 2)
        g_start = round(RNG.uniform(1.040, 1.060), 3) if name in {"kombucha"} else ""
        g_end = round(g_start - RNG.uniform(0.005, 0.020), 3) if g_start else ""
        rating = RNG.randint(2, 5)
        note_pool = {
            "kombucha":   ["tart, balanced", "too sweet, more time", "vinegary, eat as SCOBY snack",
                            "perfect fizz", "second ferment w/ ginger"],
            "koji":       ["fast bloom", "slow start, cold room", "beautiful mycelium",
                            "harvest at 48h", "missed peak"],
            "hot_sauce":  ["smoky, hot", "fermented mash too long", "added mango, great",
                            "peach variant", "fire"],
            "kimchi":     ["crisp, perfect", "salty, dilute next time", "fast bubble",
                            "added radish", "best batch yet"],
            "miso":       ["started koji + salt", "weekly stir", "first taste at 90d",
                            "dark amber", "year 1 milestone"],
        }
        notes = RNG.choice(note_pool[name])
        ws.append([f"B-{i:03d}", name, d.isoformat(), duration, vessel,
                   temp_c, start_ph, end_ph, g_start, g_end, rating, notes])
    style_sheet(ws)
    wb.save(base / "fermentation" / "batch-log.xlsx")


# ---------------------------------------------------------------------------
# CSV writers
# ---------------------------------------------------------------------------

def write_csvs(base: Path) -> None:
    # dive-sites.csv — flat version of SITES plus viz/difficulty
    dive_sites = base / "research" / "dive-sites.csv"
    dive_sites.parent.mkdir(parents=True, exist_ok=True)
    with dive_sites.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["site_id", "name", "island_group", "latitude", "longitude",
                    "typical_depth_m", "visibility_m", "difficulty", "exposure", "reef_type"])
        viz = {"Yonaguni": 30, "Ishigaki": 25, "Okinawa-honto": 18, "Kerama": 35, "Amami-oshima": 28}
        for s in SITES:
            site_id, name, group, lat, lon, depth, exp, rt = s
            v = viz[group] + RNG.randint(-5, 5)
            diff = "advanced" if depth >= 14 else "intermediate" if depth >= 9 else "beginner"
            w.writerow([site_id, name, group, lat, lon, depth, v, diff, exp, rt])

    # ph-readings.csv — 3 batches × 30 days
    ph = base / "fermentation" / "ph-readings.csv"
    ph.parent.mkdir(parents=True, exist_ok=True)
    with ph.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["timestamp", "batch_id", "ph", "temp_c"])
        for batch_id, start_ph, temp_center in [("B-007", 4.2, 24.0),
                                                  ("B-013", 5.8, 30.0),
                                                  ("B-018", 5.5, 20.0)]:
            d = date(2024, 1, 5)
            for day in range(30):
                ts = (d + timedelta(days=day)).isoformat() + "T18:00"
                # ph drifts down, temp wobbles around center
                p = round(start_ph - day * 0.03 - RNG.uniform(0, 0.15), 2)
                t = round(temp_center + RNG.uniform(-1.5, 1.5), 1)
                w.writerow([ts, batch_id, max(2.6, p), t])

    # expenses.csv — 40 Okinawa-trip expenses
    exp = base / "travel" / "okinawa-2024" / "expenses.csv"
    exp.parent.mkdir(parents=True, exist_ok=True)
    with exp.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "category", "amount_jpy", "amount_usd", "vendor", "note"])
        cats = [
            ("lodging",     [("Naha Guesthouse",   5500,  "weekly dorm"),
                             ("Sesoko Inn",        12000, "twin, 4 nights"),
                             ("Ishigaki Hostel",   6800,  "private")]),
            ("food",        [("Makishi Market",    1800,  "lunch + drinks"),
                             ("Johan's izakaya",   4500,  "dinner w/ Kenta"),
                             ("Blue Seal",         650,   "ice cream, hot day")]),
            ("dive",        [("Sea World Dive",    18000, "2-tank boat dive"),
                             ("Reef Support",      12000, "shore dive + guide"),
                             ("Umi Gitto",         22000, "charter + gear rental")]),
            ("transit",     [("JAL Naha-Ishigaki", 14000, "one-way"),
                             ("Yui Rail",          700,   "Naha to airport"),
                             ("rental scooter",    4500,  "2 days on Ishigaki")]),
            ("gear",        [("REEF safe sunscreen", 2200, "zinc-only"),
                             ("logbook",           1400,  "waterproof"),
                             ("defog",              850,  "small bottle")]),
            ("music",       [("Zoom H5 battery",   1800,  "field recording supply"),
                             ("memory card 256GB", 6500,  "spare for recordings")]),
            ("fermentation",[("koji starter",      1500,  "from local brewer"),
                             ("awamori tasting",   3500,  "3 glasses, distillery")]),
        ]
        for cat, vendors in cats:
            for vendor, jpy, note in vendors:
                d = date(2024, 3, RNG.randint(1, 30))
                usd = round(jpy / 148.0, 2)
                w.writerow([d.isoformat(), cat, jpy, usd, vendor, note])

    # finance-2024.csv — monthly summary
    fin = base / "notes" / "finance-2024.csv"
    fin.parent.mkdir(parents=True, exist_ok=True)
    with fin.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["month", "income_usd", "rent_usd", "food_usd",
                    "research_usd", "music_usd", "fermentation_usd",
                    "travel_usd", "other_usd", "savings_usd"])
        for m in range(1, 13):
            income = RNG.randint(4800, 5400)
            rent = 2200
            food = RNG.randint(450, 700)
            research = RNG.randint(50, 400)
            music = RNG.randint(0, 350)
            ferm = RNG.randint(20, 180)
            travel = 2800 if m in {3, 4} else RNG.choice([0, 0, 0, 120])
            other = RNG.randint(150, 400)
            savings = income - (rent + food + research + music + ferm + travel + other)
            w.writerow([date(2024, m, 1).strftime("%b"),
                        income, rent, food, research, music, ferm, travel, other, savings])

    # tracklist.csv — Polyp "Tidal" EP
    tl = base / "music" / "tracklist.csv"
    tl.parent.mkdir(parents=True, exist_ok=True)
    with tl.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["track_no", "title", "bpm", "key", "length_mmss",
                    "primary_samples", "notes"])
        w.writerow([1, "Tidal I (incoming)",       62, "F# minor", "04:12",
                    "Okinawa reef hydrophone, 2024-03-08",
                    "sub-bass from Korg MS-20, field recording 0:30–3:40"])
        w.writerow([2, "Polyp",                    72, "A minor",  "05:48",
                    "SCUBA inhale, dive boat hull creaks",
                    "based on lab-notebook sketch from 2023-11"])
        w.writerow([3, "Kuroshio",                 68, "C minor",  "06:24",
                    "field recording Okinawa market, 2024-03-15",
                    "rhythm from awamori distillery bottling line"])
        w.writerow([4, "Cladocopium",              80, "D minor",  "07:31",
                    "Zoom H5 hydrophone Sesoko West, 2024-04-02",
                    "dedicated to refugia site RYK-06"])
        w.writerow([5, "Tidal II ( outgoing)",     58, "F# minor", "05:09",
                    "Yonaguni drift dive, 2024-04-18",
                    "title is intentional space, see Tidal I"])
        w.writerow([6, "Six Weeks (reprise)",      66, "G minor",  "03:55",
                    "Naha airport boarding call, last day",
                    "ambient pad, closing track of EP"])


# ---------------------------------------------------------------------------
# Generated images
# ---------------------------------------------------------------------------

def write_chart(target: Path) -> None:
    """matplotlib scatter: SST vs bleaching, colored by site cluster."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    target.parent.mkdir(parents=True, exist_ok=True)
    rng = np.random.default_rng(7)

    # Three site clusters with different SST vs bleaching responses
    cluster_specs = [
        # (label, n, sst_center, sst_spread, base_bleach, slope, color)
        ("Northern refugia (RYK-05,06)",     12, 26.4, 0.7, 6.0,  1.2, "#2c7fb8"),
        ("Mid-latitude forereef (RYK-07-12)", 16, 27.6, 0.9, 25.0, 8.5, "#41b6c4"),
        ("Southern forereef (RYK-01-04)",     12, 28.4, 1.0, 45.0, 9.5, "#d7301f"),
    ]

    fig, ax = plt.subplots(figsize=(6.0, 4.0), dpi=150)
    for label, n, c, sp, b0, slope, color in cluster_specs:
        sst = rng.normal(c, sp, n)
        bleach = np.clip(b0 + (sst - c) * slope + rng.normal(0, 4, n), 0, 95)
        ax.scatter(sst, bleach, c=color, label=label, alpha=0.85,
                   edgecolors="white", linewidths=0.6, s=55)

    # Overall trend line
    all_sst = np.concatenate([rng.normal(c, sp, n) for _, n, c, sp, *_ in cluster_specs])
    all_bleach = np.concatenate([
        np.clip(b0 + (rng.normal(c, sp, n) - c) * slope + rng.normal(0, 4, n), 0, 95)
        for _, n, c, sp, b0, slope, _ in cluster_specs
    ])
    coef = np.polyfit(all_sst, all_bleach, 1)
    x_line = np.linspace(25.5, 31.0, 100)
    y_line = np.polyval(coef, x_line)
    r2 = 1 - np.sum((all_bleach - np.polyval(coef, all_sst))**2) / \
              np.sum((all_bleach - all_bleach.mean())**2)
    ax.plot(x_line, y_line, "--", color="#444", linewidth=1.0,
            label=f"linear fit (R² = {r2:.2f})")

    ax.set_xlabel("Mean summer SST (°C)")
    ax.set_ylabel("Bleaching index (%)")
    ax.set_title("Thermal forcing vs bleaching, 12 Ryukyu sites (2020–2024)",
                 fontsize=10)
    ax.set_xlim(25.5, 31.0)
    ax.set_ylim(-3, 95)
    ax.grid(True, alpha=0.25)
    ax.legend(loc="upper left", fontsize=7.5, framealpha=0.92)
    fig.tight_layout()
    fig.savefig(str(target), dpi=150, bbox_inches="tight")
    plt.close(fig)


def write_album_cover(target: Path) -> None:
    """PIL: 800x800 cover for Polyp 'Tidal' EP."""
    from PIL import Image, ImageDraw, ImageFont

    target.parent.mkdir(parents=True, exist_ok=True)
    W = H = 800
    img = Image.new("RGB", (W, H))
    px = img.load()
    # Vertical gradient navy -> teal
    top = (10, 31, 61)
    bot = (26, 107, 122)
    for y in range(H):
        r = int(top[0] + (bot[0] - top[0]) * y / H)
        g = int(top[1] + (bot[1] - top[1]) * y / H)
        b = int(top[2] + (bot[2] - top[2]) * y / H)
        for x in range(W):
            px[x, y] = (r, g, b)

    # Subtle grain (~2% noise)
    import random as _r
    _r.seed(11)
    for _ in range(int(W * H * 0.02)):
        x, y = _r.randint(0, W - 1), _r.randint(0, H - 1)
        base = px[x, y]
        d = _r.randint(-22, 22)
        px[x, y] = tuple(max(0, min(255, c + d)) for c in base)

    draw = ImageDraw.Draw(img)

    # Waveform: 60 vertical white lines, sampled from sine + noise
    rng = _r.Random(5)
    bar_count = 60
    bar_w = 6
    bar_gap = (W - bar_count * bar_w) // (bar_count + 1)
    cy = H // 2 + 80
    for i in range(bar_count):
        h = int(40 + 100 * (0.5 + 0.5 * math.sin(i * 0.42)) +
                rng.uniform(-12, 12))
        x = bar_gap + i * (bar_w + bar_gap)
        draw.rectangle([x, cy - h, x + bar_w, cy + h],
                       fill=(255, 255, 255, 230))

    # Typography
    def font(sz):
        return ImageFont.truetype("DejaVuSans-Bold.ttf", sz)
    try:
        f_title = font(74)
        f_sub = font(30)
        f_meta = font(14)
    except OSError:
        f_title = ImageFont.load_default()
        f_sub = f_title
        f_meta = f_title

    def center_text(text, font, y, fill):
        bbox = draw.textbbox((0, 0), text, font=font)
        w = bbox[2] - bbox[0]
        draw.text(((W - w) // 2, y), text, fill=fill, font=font)

    center_text("TIDAL", f_title, 180, "white")
    center_text("POLYP", f_sub, 270, (210, 220, 230))
    center_text("2024  ·  FIELD RECORDINGS", f_meta, 720, (170, 180, 190))

    img.save(str(target), "PNG", optimize=True)


# ---------------------------------------------------------------------------
# Image downloads (Picsum, with fallback)
# ---------------------------------------------------------------------------

PICSUM_TARGETS = [
    ("fermentation/photos/scoby-jar.jpg",            "scoby-jar",      800, 800),
    ("travel/okinawa-2024/photos/dive-1.jpg",        "okinawa-dive-1", 1200, 800),
    ("travel/okinawa-2024/photos/dive-2.jpg",        "okinawa-dive-2", 1200, 800),
    ("travel/okinawa-2024/photos/reef-pano.jpg",     "okinawa-reef",   1600, 600),
    ("travel/okinawa-2024/photos/market.jpg",        "okinawa-market", 1200, 800),
]


def download_one(target: Path, seed: str, w: int, h: int) -> bool:
    from PIL import Image, ImageDraw, ImageFont

    target.parent.mkdir(parents=True, exist_ok=True)
    url = f"https://picsum.photos/seed/{seed}/{w}/{h}"
    req = urllib.request.Request(url, headers={"User-Agent": "scaffold-personal/1.0"})
    last_err = None
    for attempt in (1, 2):
        try:
            with urllib.request.urlopen(req, timeout=12) as resp:
                data = resp.read()
            if len(data) < 1024:
                raise IOError(f"suspiciously small response ({len(data)} bytes)")
            target.write_bytes(data)
            print(f"  downloaded {target.name} ({len(data)//1024} KB) "
                  f"[attempt {attempt}]")
            return True
        except (urllib.error.URLError, urllib.error.HTTPError,
                IOError, TimeoutError) as e:
            last_err = e
            continue

    # Fallback: captioned placeholder
    print(f"  WARNING: download failed for {target.name} ({last_err}). "
          f"Generating captioned placeholder.")
    img = Image.new("RGB", (w, h), color=(40, 60, 80))
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("DejaVuSans.ttf", 28)
    except OSError:
        font = ImageFont.load_default()
    msg = f"{target.name}\n(download failed)\nseed={seed}"
    bbox = d.textbbox((0, 0), msg, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    d.text(((w - tw) // 2, (h - th) // 2), msg, fill="white", font=font,
           align="center")
    img.save(str(target), "JPEG", quality=85)
    return False


def download_all(base: Path) -> None:
    ok = 0
    for rel, seed, w, h in PICSUM_TARGETS:
        if download_one(base / rel, seed, w, h):
            ok += 1
    print(f"  downloaded {ok}/{len(PICSUM_TARGETS)} photos")


# ---------------------------------------------------------------------------
# CJK (Japanese + Chinese) content
# ---------------------------------------------------------------------------
# Multi-language extension. Wren studies Japanese + Mandarin and collaborates
# with Dr. Lin Wei (林伟, Xiamen University) on South China Sea corals; she also
# experiments with 红曲 (hongqu) red-yeast-rice fermentation. These writers emit
# the JA/ZH *binary* leaf files; their markdown mirrors (hand-authored under
# raw/) carry the semantic content per the leaf-node rule. The canonical facts
# below (temps, DHW breakpoint, ABV, dates, the JA-vs-EN temperature split) are
# also cited from the markdown mirrors and the eval QA, so keep them stable.

JA_FIELD_NOTES = {
    "title": "瀬底西 サンゴ礁フィールドノート",
    "subtitle": "2024年 春季（3–4月）",
    "authors": "足立 若葉（Wren Adachi）・ 田中 一兄（Hajime Tanaka）",
    "affil": "瀬底実験所・熱帯生物研究センター（琉球大学）",
    "sections": [
        ("1. 概要", [
            "本ノートは瀬底西（サイトID: RYK-06）における2024年春季調査の記録である。"
            "日本語での執筆は私（足立）の練習を兼ねており、田中研との共同データにも用いる。"
            "水温・白化・共生藻・音響の4観測を並行して実施した。"
        ]),
        ("2. 調査期間と場所", [
            "期間: 2024年3月8日〜4月2日。場所: 瀬底西、水深12〜15m、西向きの外礁斜面。"
            "水温ロガー（Onset HOBO U22-001）は水深15mに設置し、30分間隔で記録した。"
        ]),
        ("3. 水温観測", [
            "2024年3月15日、瀬底西（水深15m）の水温は 23.4°C を示した。"
            "調査期間全体では水温は 21.8〜25.0°C の範囲で推移し、春季の緩やかな上昇が観察された。"
        ]),
        ("4. 共生藻の記録", [
            "Acropora digitifera の群体において Cladocopium C3k の持続的優占を確認した"
            "（2024年3月15日、サンプル n=12）。このときの共生藻群集の多様度は Shannon H′ = 0.82 であった。"
            "C3k 優占は白化抵抗性と関連する可能性があり、英語論文の §3.3 / §3.4 と整合する。"
        ]),
        ("5. 音響観測", [
            "ハイドロフォン（Zoom H5 + Aquarian H2a）により、3月8日と4月2日に録音を行った。"
            "両日の音圧レベル（SPL）の差は 3.4 dB re 1 μPa であり、人耳でも識別可能であった。"
            "この音響データは Polyp『Tidal』EP のトラック『クラドコピウム』の素材ともなった。"
        ]),
        ("6. 特記事項", [
            "備考: 3月中旬は北西モンスーンの残響があり、視程は18m前後で安定していた。"
            "次回は共生藻の深度別構成を追加で採取する予定。"
        ]),
    ],
}

ZH_SCS_REPORT = {
    "title": "南海珊瑚白化调查报告",
    "subtitle": "2024年夏季（6–9月）",
    "authors": "林 伟（Lin Wei）· 足立 若莲（Wren Adachi）",
    "affil": "厦门大学 近海海洋环境科学国家重点实验室",
    "sections": [
        ("一、摘要", [
            "2024年夏季，我们在南海西沙群岛与海南岛周边珊瑚礁开展白化监测。"
            "白化指数与累积热应力（度·周，DHW）呈非线性关系，临界值约为 DHW = 7.8 度·周"
            "（95% 置信区间 7.3–8.3），略低于琉球群岛的报道值。"
        ]),
        ("二、调查区域与时间", [
            "区域: 南海，西沙群岛（永乐环礁、宣德环礁）及海南岛东南近岸，共6个站点。"
            "时间: 2024年6月至9月（夏季风期间），每月复查一次。"
        ]),
        ("三、水温与热应力", [
            "采用 HOBO U22-001 自记温度计，布设于水深8m，记录间隔30分钟。"
            "累积热应力依 NOAA Coral Reef Watch v3.1 方法计算。"
            "白化临界值（分段回归断点）为 DHW = 7.8 度·周（95% CI 7.3–8.3）。"
        ]),
        ("四、白化趋势", [
            "最暖站点（永乐环礁东礁）的白化比例从6月的约22%上升至9月的约58%。"
            "总体而言，2024年夏季南海北部珊瑚礁承受了中度至重度热应力。"
        ]),
        ("五、与琉球群岛的对比", [
            "与 Adachi 等（2024）报道的琉球群岛数据（断点 DHW = 8.36）相比，"
            "南海的临界值偏低，可能反映低纬度礁体对累积热应力的更高敏感性。"
            "配套数据见表 scs-bleaching-2024-zh.xlsx 与 scs-dive-sites-zh.csv。"
        ]),
        ("六、结论", [
            "南海珊瑚礁在2024年夏季呈现显著白化，需持续监测并加强区域协作。"
        ]),
    ],
}

# South China Sea survey sites (Chinese names/headers)
SCS_SITES = [
    # site_id, name_zh, region, lat, lon, depth_m, exposure_zh, reef_type_zh
    ("SCS-01", "永乐环礁东礁", "西沙群岛", 16.32, 111.65, 10, "东",   "外礁"),
    ("SCS-02", "永乐环礁西礁", "西沙群岛", 16.48, 111.41, 8,  "西",   "潟湖"),
    ("SCS-03", "宣德环礁北礁", "西沙群岛", 17.05, 111.85, 12, "北",   "外礁"),
    ("SCS-04", "海南东南近岸", "海南岛",   18.42, 110.25, 6,  "东南", "近岸礁"),
    ("SCS-05", "七连屿",       "西沙群岛", 16.95, 112.20, 9,  "东",   "外礁"),
    ("SCS-06", "亚龙湾",       "海南岛",   18.21, 109.64, 7,  "南",   "潟湖"),
]

# Ryukyu dive sites localized to Japanese (for dive-sites-ja.csv)
JA_SITE_NAMES = {
    "Yonaguni West": "与那国西", "Ishigaki South": "石垣南",
    "Kabira North": "川平北", "Sekisei Lagoon": "石西潟",
    "Cape Hedo": "辺戸岬", "Sesoko West": "瀬底西",
    "Maeda Point": "真喜田ポイント", "Onna Village East": "恩納村東",
    "Kerama Inner": "慶良間内", "Kerama Outer": "慶良間外",
    "Amami North": "奄美北", "Amami South": "奄美南",
}
JA_GROUP = {"Yonaguni": "与那国島", "Ishigaki": "石垣島",
            "Okinawa-honto": "沖縄本島", "Kerama": "慶良間諸島",
            "Amami-oshima": "奄美大島"}
JA_EXPOSURE = {"N": "北", "S": "南", "E": "東", "W": "西",
               "NE": "北東", "NW": "北西", "SE": "南東", "SW": "南西"}
JA_REEF = {"forereef": "外礁", "backreef": "裏礁", "lagoon": "潟湖"}

# Japanese-release tracklist for the Tidal EP (kanji title 潮 = tide)
JA_TRACKLIST = [
    # track_no, title_ja, bpm, key, length_mmss, primary_samples, notes
    [1, "潮 I（差し潮）",   62, "F#短調", "04:12",
     "沖縄サンゴ礁ハイドロフォン、2024-03-08", "Korg MS-20 サブベース、フィールド録音 0:30–3:40"],
    [2, "ポリプ",           72, "イ短調", "05:48",
     "SCUBA 吸気、ダイブ船の軋み", "2023-11 のスケッチが基"],
    [3, "黒潮",             68, "ハ短調", "06:24",
     "沖縄市場の環境音、2024-03-15", "泡盛蒸留所の瓶詰めラインのリズム"],
    [4, "クラドコピウム",   80, "ニ短調", "07:31",
     "Zoom H5 ハイドロフォン 瀬底西、2024-04-02", "避難礁サイト RYK-06 に捧ぐ"],
    [5, "潮 II（引き潮）",  58, "F#短調", "05:09",
     "与那国ドリフトダイブ、2024-04-18", "タイトルの空白は意図的（潮 I 参照）"],
    [6, "六週間（再び）",   66, "ト短調", "03:55",
     "那覇空港搭乗案内、最終日", "アンビエント・パッド、EP の終曲"],
]

# Awamori mash batches (泡盛, black-koji / 黒麹). Row A-003 carries the
# canonical final gravity / ABV cited by the eval QA (Q34).
AWAMORI_ROWS = [
    # batch_id, type, start_date, days, vessel, temp_c, sg_start, sg_end, abv, rating, notes
    ["A-001", "泡盛（米・黒麹）", "2024-02-20", 16, "甕（かめ）",   30, 1.052, 0.990, 38.0, 4, "田中研の米を使用"],
    ["A-002", "泡盛（米・黒麹）", "2024-03-02", 15, "ステンレスタンク", 31, 1.050, 0.988, 40.5, 4, "比嘉さん指導"],
    ["A-003", "泡盛（米・黒麹）", "2024-03-10", 14, "甕（かめ）",   32, 1.054, 0.984, 43.2, 5, "最高の出来、蒸留済み"],
    ["A-004", "泡盛（米・黒麹）", "2024-03-22", 14, "ステンレスタンク", 31, 1.048, 0.986, 41.0, 3, "発酵やや遅れ"],
]

CJK_PICSUM = [
    ("research/photos/scs-reef-zh.jpg",        "scs-coral-reef",  1200, 800),
    ("fermentation/photos/hongqu-rice-zh.jpg", "hongqu-red-rice", 800, 800),
]


def _register_cjk(font_name: str) -> None:
    """Register a reportlab bundled Adobe CJK CID font (no external files)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    try:
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))
    except Exception:  # already registered or missing — ignore
        pass


def write_pdf_ja(target: Path) -> None:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

    target.parent.mkdir(parents=True, exist_ok=True)
    font = "HeiseiMin-W3"  # Japanese Mincho, bundled with reportlab
    _register_cjk(font)
    base = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(target), pagesize=LETTER,
        leftMargin=1.0 * inch, rightMargin=1.0 * inch,
        topMargin=1.0 * inch, bottomMargin=1.0 * inch,
        title=JA_FIELD_NOTES["title"], author=JA_FIELD_NOTES["authors"])
    title_s = ParagraphStyle("JaTitle", parent=base["Title"], fontName=font,
        fontSize=20, leading=26, spaceAfter=10)
    h1 = ParagraphStyle("JaH1", parent=base["Heading1"], fontName=font,
        fontSize=13, leading=18, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("JaBody", fontName=font, fontSize=11,
        leading=17, spaceAfter=8, alignment=4)
    small = ParagraphStyle("JaSmall", fontName=font, fontSize=9, leading=13)

    flow = [Paragraph(JA_FIELD_NOTES["title"], title_s),
            Paragraph(JA_FIELD_NOTES["subtitle"], body),
            Paragraph(JA_FIELD_NOTES["authors"], body),
            Paragraph(JA_FIELD_NOTES["affil"], small), Spacer(1, 8)]
    for heading, paras in JA_FIELD_NOTES["sections"]:
        flow.append(Paragraph(heading, h1))
        for p in paras:
            flow.append(Paragraph(p, body))
        flow.append(Spacer(1, 4))
    doc.build(flow)


def write_pdf_zh(target: Path) -> None:
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

    target.parent.mkdir(parents=True, exist_ok=True)
    font = "STSong-Light"  # Simplified Chinese Song, bundled with reportlab
    _register_cjk(font)
    base = getSampleStyleSheet()
    doc = SimpleDocTemplate(str(target), pagesize=LETTER,
        leftMargin=1.0 * inch, rightMargin=1.0 * inch,
        topMargin=1.0 * inch, bottomMargin=1.0 * inch,
        title=ZH_SCS_REPORT["title"], author=ZH_SCS_REPORT["authors"])
    title_s = ParagraphStyle("ZhTitle", parent=base["Title"], fontName=font,
        fontSize=20, leading=26, spaceAfter=10)
    h1 = ParagraphStyle("ZhH1", parent=base["Heading1"], fontName=font,
        fontSize=13, leading=18, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("ZhBody", fontName=font, fontSize=11,
        leading=17, spaceAfter=8, alignment=4)
    small = ParagraphStyle("ZhSmall", fontName=font, fontSize=9, leading=13)

    flow = [Paragraph(ZH_SCS_REPORT["title"], title_s),
            Paragraph(ZH_SCS_REPORT["subtitle"], body),
            Paragraph(ZH_SCS_REPORT["authors"], body),
            Paragraph(ZH_SCS_REPORT["affil"], small), Spacer(1, 8)]
    for heading, paras in ZH_SCS_REPORT["sections"]:
        flow.append(Paragraph(heading, h1))
        for p in paras:
            flow.append(Paragraph(p, body))
        flow.append(Spacer(1, 4))
    doc.build(flow)


def write_excel_zh(base: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    (base / "research").mkdir(parents=True, exist_ok=True)
    header_fill = PatternFill("solid", fgColor="1A3B5C")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    def style_sheet(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 32)

    wb = Workbook()
    ws = wb.active
    ws.title = "站点"  # Sites
    ws.append(["站点编号", "名称", "区域", "纬度", "经度", "水深(米)", "朝向", "礁型"])
    for s in SCS_SITES:
        ws.append(list(s))
    style_sheet(ws)

    ws = wb.create_sheet("月度水温")  # SST monthly
    ws.append(["月份"] + [s[0] for s in SCS_SITES])
    months_zh = ["1月", "2月", "3月", "4月", "5月", "6月",
                 "7月", "8月", "9月", "10月", "11月", "12月"]
    scs_base = [22.0, 22.4, 24.0, 26.2, 28.3, 29.6, 30.2, 30.4, 29.5, 27.6, 25.1, 22.8]
    for i, m in enumerate(months_zh):
        row = [m]
        for s in SCS_SITES:
            row.append(round(scs_base[i] + (18.0 - s[3]) * -0.6 + RNG.uniform(-0.3, 0.3), 2))
        ws.append(row)
    style_sheet(ws)

    ws = wb.create_sheet("季度白化百分比")  # Bleaching quarterly, 2024
    ws.append(["季度"] + [s[0] for s in SCS_SITES])
    for q in (1, 2, 3, 4):
        row = [f"2024Q{q}"]
        for s in SCS_SITES:
            # Summer (Q2/Q3) bleaching much higher; SCS-01 warmest
            base_bleach = {1: 6, 2: 22, 3: 58, 4: 19}[q]
            adj = -6 if s[0] in {"SCS-02", "SCS-06"} else (8 if s[0] == "SCS-01" else 0)
            row.append(round(max(0, base_bleach + adj + RNG.uniform(-4, 4)), 1))
        ws.append(row)
    style_sheet(ws)

    wb.save(base / "research" / "scs-bleaching-2024-zh.xlsx")


def write_csvs_cjk(base: Path) -> None:
    # research/dive-sites-ja.csv — Ryukyu sites localized to Japanese
    ja = base / "research" / "dive-sites-ja.csv"
    ja.parent.mkdir(parents=True, exist_ok=True)
    viz = {"Yonaguni": 30, "Ishigaki": 25, "Okinawa-honto": 18, "Kerama": 35, "Amami-oshima": 28}
    with ja.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["サイトID", "名称", "島嶼群", "緯度", "経度",
                    "典型深度(m)", "視程(m)", "難易度", "朝向", "礁型"])
        for s in SITES:
            site_id, name, group, lat, lon, depth, exp, rt = s
            v = viz[group] + RNG.randint(-5, 5)
            diff = "上級" if depth >= 14 else "中級" if depth >= 9 else "初級"
            w.writerow([site_id, JA_SITE_NAMES.get(name, name), JA_GROUP[group],
                        lat, lon, depth, v, diff, JA_EXPOSURE.get(exp, exp), JA_REEF[rt]])

    # research/scs-dive-sites-zh.csv — South China Sea sites (Chinese)
    zh = base / "research" / "scs-dive-sites-zh.csv"
    with zh.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["站点编号", "名称", "区域", "纬度", "経度",
                    "典型深度(米)", "能见度(米)", "难度", "朝向", "礁型"])
        diff_zh = {10: "中级", 8: "初级", 12: "高级", 6: "初级", 9: "中级", 7: "初级"}
        viz_by_region = {"西沙群岛": 30, "海南岛": 18}
        for site_id, name, region, lat, lon, depth, exp, rt in SCS_SITES:
            v = viz_by_region[region] + RNG.randint(-4, 4)
            w.writerow([site_id, name, region, lat, lon, depth, v,
                        diff_zh.get(depth, "中级"), exp, rt])

    # fermentation/hongqu-ph-readings-zh.csv — 红曲 fermentation, warm (~33°C)
    hq = base / "fermentation" / "hongqu-ph-readings-zh.csv"
    hq.parent.mkdir(parents=True, exist_ok=True)
    with hq.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["时间", "批次编号", "pH", "温度(°C)"])
        d = date(2024, 7, 5)
        for day in range(14):
            ts = (d + timedelta(days=day)).isoformat() + "T20:00"
            p = round(6.0 - day * 0.11 - RNG.uniform(0, 0.08), 2)
            t = round(33.0 + RNG.uniform(-1.2, 1.5), 1)
            w.writerow([ts, "HQ-002", max(4.4, p), t])

    # fermentation/awamori-batch-ja.csv — 泡盛 mash log (Japanese)
    aw = base / "fermentation" / "awamori-batch-ja.csv"
    with aw.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["バッチID", "種類", "開始日", "日数", "タンク",
                    "温度(°C)", "開始比重", "終了比重", "ABV(%)", "評価", "備考"])
        for r in AWAMORI_ROWS:
            w.writerow(r)

    # music/tracklist-ja.csv — Japanese-release Tidal EP tracklist
    tl = base / "music" / "tracklist-ja.csv"
    tl.parent.mkdir(parents=True, exist_ok=True)
    with tl.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["トラック番号", "タイトル", "BPM", "キー", "長さ",
                    "主なサンプル", "備考"])
        for r in JA_TRACKLIST:
            w.writerow(r)


def _cover_fallback_pil(target: Path) -> None:
    """Minimal PIL cover used only if CJK rasterization is unavailable. Renders
    'TIDAL' (the kanji title 潮 is documented in tidal-ep-liner-notes-ja.md)."""
    from PIL import Image, ImageDraw, ImageFont
    W = H = 800
    img = Image.new("RGB", (W, H), (61, 10, 31))
    draw = ImageDraw.Draw(img)
    try:
        f_title = ImageFont.truetype("DejaVuSans-Bold.ttf", 74)
        f_sub = ImageFont.truetype("DejaVuSans-Bold.ttf", 30)
    except OSError:
        f_title = f_sub = ImageFont.load_default()

    def center(text, font, y, fill):
        bbox = draw.textbbox((0, 0), text, font=font)
        draw.text(((W - (bbox[2] - bbox[0])) // 2, y), text, fill=fill, font=font)

    center("TIDAL", f_title, 180, "white")
    center("POLYP", f_sub, 330, (240, 220, 220))
    center("2024  ·  JP EDITION", f_sub, 720, (200, 180, 190))
    img.save(str(target), "PNG", optimize=True)


def write_album_cover_ja(target: Path) -> None:
    """800x800 Japanese-release cover with kanji title 潮 (tide).

    Built as a reportlab PDF (bundled CJK CID font always renders), then
    rasterized to PNG with PyMuPDF. Falls back to a PIL 'TIDAL' cover if
    PyMuPDF is unavailable — in that case the kanji title 潮 is still documented
    in the markdown mirror tidal-ep-liner-notes-ja.md."""
    import os
    import tempfile
    import random as _r
    from reportlab.pdfgen import canvas

    target.parent.mkdir(parents=True, exist_ok=True)
    W = H = 800
    font = "HeiseiMin-W3"
    _register_cjk(font)

    pdf_path = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False).name
    c = canvas.Canvas(pdf_path, pagesize=(W, H))
    # Gradient background: crimson (bottom) -> magenta (top).
    bands = 64
    for i in range(bands):
        frac = i / (bands - 1)
        c.setFillColorRGB((61 + int((122 - 61) * frac)) / 255,
                          (10 + int((26 - 10) * frac)) / 255,
                          (31 + int((107 - 31) * frac)) / 255)
        c.rect(0, i * (H / bands), W, H / bands + 1, fill=1, stroke=0)
    # Waveform bars.
    rng = _r.Random(9)
    bar_count, bar_w = 60, 6
    bar_gap = (W - bar_count * bar_w) // (bar_count + 1)
    cy = H // 2 - 80
    c.setFillColorRGB(1, 1, 1)
    for i in range(bar_count):
        h = int(40 + 100 * (0.5 + 0.5 * math.sin(i * 0.42)) + rng.uniform(-12, 12))
        x = bar_gap + i * (bar_w + bar_gap)
        c.rect(x, cy - h, bar_w, 2 * h, fill=1, stroke=0)
    # Typography.
    c.setFillColorRGB(1, 1, 1)
    c.setFont(font, 150)
    c.drawCentredString(W / 2, H - 300, "潮")
    c.setFont(font, 34)
    c.setFillColorRGB(240 / 255, 220 / 255, 220 / 255)
    c.drawCentredString(W / 2, H - 360, "POLYP")
    c.setFont(font, 16)
    c.setFillColorRGB(200 / 255, 180 / 255, 190 / 255)
    c.drawCentredString(W / 2, 40, "2024  ·  日本盤  ·  FIELD RECORDINGS")
    c.showPage()
    c.save()

    try:
        import fitz  # PyMuPDF — self-contained wheel, no system deps
        doc = fitz.open(pdf_path)
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1, 1))  # 800pt -> 800px
        pix.save(str(target))
        doc.close()
    except Exception as e:
        print(f"  WARNING: CJK rasterize failed ({e}); JP cover fell back to 'TIDAL'.")
        _cover_fallback_pil(target)
    finally:
        os.unlink(pdf_path)


def write_cjk_all(base: Path) -> None:
    write_pdf_ja(base / "research" / "sesoko-field-notes-ja.pdf")
    write_pdf_zh(base / "research" / "scs-bleaching-survey-zh.pdf")
    write_excel_zh(base)
    write_csvs_cjk(base)
    write_album_cover_ja(base / "music" / "tidal-ep-cover-ja.png")
    ok = sum(download_one(base / rel, seed, w, h)
             for rel, seed, w, h in CJK_PICSUM)
    print(f"  downloaded {ok}/{len(CJK_PICSUM)} CJK photos")


# ---------------------------------------------------------------------------
# Driver
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                    help="print planned file list without writing")
    ap.add_argument("--only",
                    choices=["pdf", "excel", "csv", "images", "download", "cjk"],
                    help="run only one section")
    args = ap.parse_args()

    plan = [
        ("PDF",     P / "research" / "coral-resilience-paper.pdf"),
        ("Excel",   P / "research" / "bleaching-survey-2024.xlsx"),
        ("Excel",   P / "fermentation" / "batch-log.xlsx"),
        ("CSV",     P / "research" / "dive-sites.csv"),
        ("CSV",     P / "fermentation" / "ph-readings.csv"),
        ("CSV",     P / "travel" / "okinawa-2024" / "expenses.csv"),
        ("CSV",     P / "notes" / "finance-2024.csv"),
        ("CSV",     P / "music" / "tracklist.csv"),
        ("Image",   P / "research" / "figures" / "temp-vs-bleach.png"),
        ("Image",   P / "music" / "album-cover.png"),
        ("Photo",   P / "fermentation" / "photos" / "scoby-jar.jpg"),
        ("Photo",   P / "travel" / "okinawa-2024" / "photos" / "dive-1.jpg"),
        ("Photo",   P / "travel" / "okinawa-2024" / "photos" / "dive-2.jpg"),
        ("Photo",   P / "travel" / "okinawa-2024" / "photos" / "reef-pano.jpg"),
        ("Photo",   P / "travel" / "okinawa-2024" / "photos" / "market.jpg"),
        # --- CJK (Japanese + Chinese) leaf files ---
        ("PDF",     P / "research" / "sesoko-field-notes-ja.pdf"),
        ("PDF",     P / "research" / "scs-bleaching-survey-zh.pdf"),
        ("Excel",   P / "research" / "scs-bleaching-2024-zh.xlsx"),
        ("CSV",     P / "research" / "dive-sites-ja.csv"),
        ("CSV",     P / "research" / "scs-dive-sites-zh.csv"),
        ("CSV",     P / "fermentation" / "hongqu-ph-readings-zh.csv"),
        ("CSV",     P / "fermentation" / "awamori-batch-ja.csv"),
        ("CSV",     P / "music" / "tracklist-ja.csv"),
        ("Image",   P / "music" / "tidal-ep-cover-ja.png"),
        ("Photo",   P / "research" / "photos" / "scs-reef-zh.jpg"),
        ("Photo",   P / "fermentation" / "photos" / "hongqu-rice-zh.jpg"),
    ]
    print(f"Scaffolding under: {P}")
    if args.dry_run:
        print(f"  planned: {len(plan)} files")
        for kind, p in plan:
            print(f"    [{kind:6}] {p.relative_to(ROOT)}")
        return 0

    P.mkdir(parents=True, exist_ok=True)
    if args.only in (None, "pdf"):
        print("Writing PDF...")
        write_pdf(P / "research" / "coral-resilience-paper.pdf")
    if args.only in (None, "excel"):
        print("Writing Excels...")
        write_excels(P)
    if args.only in (None, "csv"):
        print("Writing CSVs...")
        write_csvs(P)
    if args.only in (None, "images"):
        print("Writing generated images...")
        write_chart(P / "research" / "figures" / "temp-vs-bleach.png")
        write_album_cover(P / "music" / "album-cover.png")
    if args.only in (None, "download"):
        print("Downloading photos...")
        download_all(P)
    if args.only in (None, "cjk"):
        print("Writing CJK (JA/ZH) files...")
        write_cjk_all(P)
    print(f"Done. {len(plan)} files targeted.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
